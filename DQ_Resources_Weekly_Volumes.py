import os
import pandas as pd
import time
from datetime import datetime
import openpyxl

start_time = time.time()

# Specify the paths to the folders containing the excel files
folder_paths = [
    r'path1',
    r'path2',
    r'path3',
]

# Define the output folder path
output_folder_path = r"output_path"

# Create output folder if not exists
if not os.path.exists(output_folder_path):
    os.makedirs(output_folder_path)
    print(f"Created output folder at: {output_folder_path}")

# List to store data frames
dfs = []

# Loop through each folder path
for folder_path in folder_paths:

    # Check if folder path exists
    if not os.path.exists(folder_path):
        print(f"Folder path '{folder_path}' does not exist. Skipping...")
        continue

    print(f"Processing files in folder: {folder_path}")

    # Loop through each file in the folder
    for filename in os.listdir(folder_path):
        
        # Check if the file ends with .xlsx and is not a temporary file
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            
            # Define the full path to the file
            file_path = os.path.join(folder_path, filename)
            
            # Try reading the file
            try:
                df = pd.read_excel(file_path)
                
                # Check if necessary columns are in the file
                if set(["Formula", "Resource name", "Date", "Month"]).issubset(df.columns):
                    
                    # Convert 'Month' column to datetime for comparison, assuming 'Month' as 'MMM-YY' format
                    df['Month'] = pd.to_datetime(df['Month'], format='%b-%y')
                    
                    # Keep only rows where 'Month' is on or after January 2023 and necessary columns are not null
                    comparison_date = datetime(2023, 1, 1)  # January 2023
                    df = df[df['Month'] >= comparison_date]
                    df = df[df[["Formula", "Resource name", "Date", "Month"]].notnull().all(axis=1)]
                    
                    if not df.empty:
                        # Dynamically adjust 'Month' column to reflect the month and year correctly
                        # Format 'Month' column to display as 'MMM-YY'
                        df['Month'] = df['Month'].dt.strftime('%b-%y')
                        
                        dfs.append(df)
                        print(f"Added data from file: {filename}")
                    else:
                        print(f"File {filename} does not contain valid data. Skipping...")
            
            except Exception as e:
                print(f"Could not read file {filename}. Error: {e}")

# After concatenating all data frames
if dfs:
    result_df = pd.concat(dfs, ignore_index=True)
    
    # Convert 'Date' column to datetime format if it's not already
    result_df['Date'] = pd.to_datetime(result_df['Date'])
    result_df['Actual Date'] = pd.to_datetime(result_df['Actual Date'])

    # Ensure the 'Month' column is in datetime format before setting the day
    result_df['Month'] = pd.to_datetime(result_df['Month'], format='%b-%y')
    result_df['Month'] = result_df['Month'].apply(lambda x: x.replace(day=24))

    # Transpose specified columns
    result_df = result_df.melt(id_vars=["Formula", "Resource name", "Date", "Month", "Category", "Work Drivers", "Activity", "Asset Class", "Case #", "Error Count", "Actual Date", "ID number"],
                               value_vars=["Count", "Setup", "Amend", "Review", "Closure", "4 eye Count"],
                               var_name='Name',
                               value_name='Value')

    # Add the 'Source' column with all values set to "Orchestra"
    result_df['Source'] = "Orchestra"

    # Add the 'InAccuracy' column with all values set to "Accurate"
    result_df['InAccuracy'] = "Accurate"

    # Temporarily change 'Month' to string to prevent to_excel from changing the format
    result_df['Month'] = result_df['Month'].dt.strftime('%m/%d/%Y')

    # Ensure the columns are in the specified order
    result_df = result_df[["Formula", "Resource Name", "Date", "Month", "Category", "Work Drivers", "Activity", "Asset Class", "Case #", "Error Count", "Actual Date", "ID number", "Source", "Name", "Value", "InAccuracy"]]

    # Save the concatenated data frame to a new Excel file
    output_file_path = os.path.join(output_folder_path, "Resources_Daily_Volumes_Data.xlsx")
    result_df.to_excel(output_file_path, index=False)

    # Apply custom formatting for the 'Month' column using openpyxl
    wb = openpyxl.load_workbook(output_file_path)
    ws = wb.active

    # Apply the custom date format for the 'Month' column, assumed to be column D
    for cell in ws['C'][1:]:  # Skip the header row
        cell.number_format = 'm/d/yyyy'

    # Apply the custom date format for the 'Month' column, assumed to be column D
    for cell in ws['D'][1:]:  # Skip the header row
        if cell.value:  # Check if cell is not empty
            cell.value = datetime.strptime(cell.value, '%m/%d/%Y')  # Convert back to datetime
            cell.number_format = 'dd-mmm'

    # Apply the custom date format for the 'Actual Date' column, assumed to be column Q
    for cell in ws['Q'][1:]:  # Skip the header row
        cell.number_format = 'm/d/yyyy'

    wb.save(output_file_path)

    print(f"Data concatenated successfully. Output file saved at: {output_folder_path}")
else:
    print("No valid data found. Concatenation skipped.")

end_time = time.time()
execution_time = (end_time - start_time) / 60
print(f"Script execution completed in {execution_time:.2f} minutes")